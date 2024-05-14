#cadastro de clientes com o custom Tkinter !!
#bibliotecas usadas 
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl
import pathlib
from openpyxl import Workbook

#assim se faz o set de aparencia padrão do Sistema:
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")
#A classe super é a classe principal do app

#para fazermos as definições/ teremos de chama-los abaixo do super (dentro do def __init__)
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appeareance()
        self.all_system()
        

#geometria da janela que irá abrir
    def layout_config(self):
        self.title("Sistema gestor de clientes")
        self.geometry("700x500")
        
#______________________________________________________________________________________________
        
#aparencia da janela
    def appeareance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent",text_color=['#000','#fff']).place (x=50,y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["System","Dark","Light",],command=self.change_apm).place (x=50,y=460)

#_______________________________________________________________________________________________
        
#Fontes e escritas da janela 
    def all_system(self):
        frame = ctk.CTkFrame(self, width=700,height=50,corner_radius=0,bg_color="teal",fg_color="teal").place(x=0,y=10)

        title = ctk.CTkLabel(frame, text="Sistema gestor de clientes", font=("Century Gothic bold",24),text_color="#fff",bg_color="teal").place(x=215,y=20)

        span = ctk.CTkLabel(frame, text="Por favor certifique-se de preencher todos os campos!",font=("Century Gothic bold",16),text_color=["#000","#fff"]).place(x=50, y=90)

#____________________________________________________________________________________________________________________________________________________________________________

#Criação da planilha e da folha no excel:
        planilha= pathlib.Path("Clientes.xlsx")
        if planilha.exists():
             pass
        else:
            planilha=Workbook()
            folha=planilha.active
            folha['A1']="Nome completo"
            folha['B1']="CPF"
            folha['C1']="Idade"
            folha['D1']="Genero"
            folha['E1']="Endereço"
            folha['F1']="Observações"

            planilha.save('Clientes.xlsx')

#__________________________________________________________________________________________            
#coletando os dados das entradas:

        def submit():
            name=name_value.get()
            cpf=cpf_value.get()
            age=age_value.get()
            gender=gender_combobox.get()
            address=address_value.get()
#pegando os dados da caixa de texto
            obs=obs_entry.get(0.0,END)
            

#____________________________________________________________________________________________

#tratamento de erro caso os campos não estejam preenchidos!!!
            if(name=="" or cpf=="" or age=="" or address==""):
                messagebox.showerror("ATENÇÃO","Erro\nPor favor preencha todos os campos!")

#_____________________________________________________________________________________________
            else:
#colocando dados na planilha
                planilha = openpyxl.load_workbook('Clientes.xlsx')
                folha=planilha.active

                cpf=cpf.replace(".","").replace("-","")

#Verificando se o CPF ja está cadastrado

                cpf_existente = False
                for row in folha.iter_rows(min_row=2,max_row = folha.max_row ,min_col=2,max_col=2):
                    if row[0].value == cpf:
                        cpf_existente = True
                        break
                if cpf_existente:
                    messagebox.showerror("ERRO","CPF cadastrado anteriormente!")

                else:
                        max_row = folha.max_row + 1
                        folha.cell(column=1,row=folha.max_row+1,value=name)
                        folha.cell(column=2,row=folha.max_row,value=cpf)
                        folha.cell(column=3,row=folha.max_row,value=age)
                        folha.cell(column=4,row=folha.max_row,value=gender)
                        folha.cell(column=5,row=folha.max_row,value=address)
                        folha.cell(column=6,row=folha.max_row,value=obs)

                        planilha.save(r"Clientes.xlsx")
                        messagebox.showinfo("Sistema","Dados salvos com sucesso!")

#DEF para podermos ver os dados da planilha sem acessa-la no exel

        def show_data():
            planilha = openpyxl.load_workbook('Clientes.xlsx')
            folha = planilha.active

            data_window = Toplevel(self)
            data_window.title("Dados dos Clientes")
            data_window.geometry("600x400")

            for i, row in enumerate(folha.iter_rows(values_only=True)):
                for j, value in enumerate(row):
                    label = Label(data_window, text=value)
                    label.grid(row=i, column=j)
#______________________________________________________________________________________________

#Botões para limpar os campos preenchidos:
        def clear():
            name_value.set("")
            cpf_value.set("")
            age_value.set("")
            address_value.set("")
#apagando os dados da caixa de texto
            obs_entry.delete(0.0,END)
            pass

#________________________________________________________________________________________________

#variaveis de texto:

        name_value = StringVar()
        cpf_value=StringVar()
        age_value=StringVar()
        address_value=StringVar()

#_________________________________________________________________________________________________        

#todas as entradas:
        name_entry = ctk.CTkEntry(self, width= 350, textvariable= name_value, font=("Century Gothic bold",16),fg_color="transparent")
        cpf_entry = ctk.CTkEntry(self, width= 200,textvariable=cpf_value, font=("Century Gothic bold",16),fg_color="transparent")
        age_entry = ctk.CTkEntry(self, width= 150,textvariable=age_value, font=("Century Gothic bold",16),fg_color="transparent")
        address_entry = ctk.CTkEntry(self, width= 200,textvariable=address_value, font=("Century Gothic bold",16),fg_color="transparent")

#Agora é usado o combobox para posicionar a variavél genero(gender)
        gender_combobox = ctk.CTkComboBox(self,values=["Masculino","Feminino","Outro"], font= ("Century Gothic bold",14),width=150)
        gender_combobox.set("Masculino")

#Entrada das observações (obs):
        obs_entry=ctk.CTkTextbox(self,width=500,height=150,font=("arial",18),border_color="#aaa",border_width=2,fg_color="transparent")

#____________________________________________________________________________________________________________________________________________

#Todas as labels
        lb_name = ctk.CTkLabel(frame, text="Nome completo:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        
        lb_cpf = ctk.CTkLabel(frame, text="CPF:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        
        lb_age = ctk.CTkLabel(frame, text="Idade:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        
        lb_gender = ctk.CTkLabel(frame, text="Genero:",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        
        lb_address = ctk.CTkLabel(frame, text="Endereço:", font=("Century Gothic bold",16),text_color=["#000","#fff"])
        
        lb_obs = ctk.CTkLabel(frame, text="Observações: ",font=("Century Gothic bold",16),text_color=["#000","#fff"])
        
        btn_submit = ctk.CTkButton(self,text="Salvar dados".upper(),command=submit,fg_color="#151",hover_color="#175").place(x=300,y=420)

        btn_submit = ctk.CTkButton(self,text="Limpar os campos".upper(),command=clear,fg_color="#178",hover_color="#175").place(x=500,y=420)

        btn_show_data = ctk.CTkButton(self,text="Mostrar Dados".upper(),command=show_data,fg_color="#178",hover_color="#175").place(x=50,y=420)
#_______________________________________________________________________________________________________________________________________________

#posicionando as variavéis da janela:
#na parte do "var_entry" contamos a diferença de 30 pixels no eixo y
#por ex:o "lb_var.place" está configurado assim: x=50,y=120 
#então o "var_entry" irá ficar x=50,y=150, assim da uma boa distancia do campo de entrada e do texto do app
#A unica exeção é o obs afinal ele é um campo grande de inserção de texto
        lb_name.place(x=50,y=120)
        name_entry.place(x=50,y=150)

        lb_cpf.place(x=450,y=120)
        cpf_entry.place(x=450,y=150)

        lb_age.place(x=300,y=190)
        age_entry.place(x=300,y=220)

        lb_gender.place(x=500,y=190)
        gender_combobox.place(x=500,y=220)

        lb_address.place(x=50,y=190)
        address_entry.place(x=50,y=220)

        lb_obs.place(x=50,y=260)
        obs_entry.place(x=180,y=260)

#__________________________________________________________________________________________________________________________

#Definição de nova aparencia na hora da seleção pelo botão
    def change_apm(self, new_appearence_mode):
        ctk.set_appearance_mode(new_appearence_mode)

#____________________________________________________________________________________________________________________________

#Loop para a janela do app se manter ativa
if __name__ == "__main__":
    app=App()
    app.resizable(width=False,height=False)
    app.mainloop()

#______________________________________________________________________________________________________________________________

